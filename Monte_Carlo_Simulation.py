import random
import math
from statistics import mean
import matplotlib
matplotlib.use('Agg')  # Non-GUI backend for matplotlib
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

import Marbles_Drop_Simulation as MarblesPyFile


def Monte_Carlo_Simulation():
    experiments_count = int(input("How many experiments do you need to run each time? "))
    show_graph = input("Do you want to display the simulation graph? (y/n): ").lower() == 'y'
    save_to_excel_flag = input("Do you want to save the results to an existing Excel file? (y/n): ").lower() == 'y'
    save_marble_dropping_image = input("Would you like to save the marble dropping areas image from the simulation? (y/n): ").lower() == 'y'


    if save_marble_dropping_image:
        MarblesPyFile.mcs_MarblesDropSimulation()
    
    sample_type_list = [1000, 10000, 100000, 1000000]

    # Run the simulation and log the results
    pi_results, probability_list = run_simulation_and_log(sample_type_list, experiments_count)

    # Calculate statistics and plot the graph
    calculate_statistics_and_plot(pi_results, show_graph)

    # Save the results to an Excel file
    if save_to_excel_flag:
        file_path = "./Coursework.xlsx"
        sheet_name = "Monte Carlo Simulation"
        update_excel_file(file_path, sheet_name, sample_type_list, pi_results, probability_list)
    
    
# Run the simulation and log the results
def run_simulation_and_log(sample_type_list, experiments_count):
    pi_results = {count: [] for count in sample_type_list}
    probability_list = []

    for sample_type in sample_type_list:
        print(f"\nRunning {experiments_count} experiments for {sample_type} trials...")
        for experiment in range(experiments_count):
            pi_estimate, prob_circle, prob_square, prob_union= drop_marbles(sample_type)

            probability_record = {
            "Round": experiment + 1,
            "Trial Count": sample_type,
            "Probability Circle": prob_circle,
            "Probability Square": prob_square,
            "Probability Union": prob_union
            }
            probability_list.append(probability_record)
            pi_results[sample_type].append(pi_estimate)
            print(f"Experiment {experiment + 1}: Estimated Pi = {pi_estimate:.6f}")

    return pi_results, probability_list

# Function to drop marbles and estimate Pi
def drop_marbles(num_trials):
    circle_hits = 0
    square_hits = 0

    for _ in range(num_trials):
        x = random.uniform(-2, 4)
        y = random.uniform(-2, 2)

        if (2 < x < 3 and -0.5 < y < 0.5):
            square_hits += 1
        elif x**2 + y**2 <= 1:
            circle_hits += 1

    estimated_pi = circle_hits / square_hits if square_hits != 0 else 0
    prob_circle = circle_hits / num_trials
    prob_square = square_hits / num_trials
    prob_union = (circle_hits + square_hits) / num_trials

    return estimated_pi, prob_circle, prob_square, prob_union

# Calculate statistics and plot the graph
def calculate_statistics_and_plot(pi_results, show_graph=True):
    trial_counts = list(pi_results.keys())
    means = []    
    
    print("\nStatistical Summary:")
    for num_trials, estimates in pi_results.items():
        mean_pi = mean(estimates)        
        means.append(mean_pi)      
        
        print(f"For N = {num_trials}: Mean Pi = {mean_pi:.6f}")
    
    if show_graph:
        plt.figure(figsize=(10, 6))
        plt.plot([str(tc) for tc in trial_counts], means, marker='o', color='b', label='Mean Pi')
        plt.axhline(y=math.pi, color='r', linestyle='--', label="Actual Pi")
        plt.title("Estimated Pi vs. Number of Trials")
        plt.xlabel("Number of Trials (N)")
        plt.ylabel("Estimated Pi")
        plt.legend()
        plt.grid(True)
        plt.savefig('pi_estimate_plot.png')
        print("Plot saved as 'pi_estimate_plot.png'.")


# Update the Excel file with the simulation results
def update_excel_file(file_path, sheet_name, trial_counts, pi_results, probability_list):

    book = load_workbook(file_path)
    if sheet_name not in book.sheetnames:
        print(f"Sheet '{sheet_name}' not found.")
        return

    sheet = book[sheet_name]

    # Clear existing data in the table
    for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row, min_col=2, max_col=6):
        for cell in row:
            cell.value = None

    # Clear previous data for mean and mode in I7:L8 (Mean in row 7, Mode in row 8)
    for row in range(7, 9):  
        for col in range(9, 13):  
            sheet.cell(row=row, column=col).value = None

    # Fill in trial numbers in column B
    trial_numbers = list(range(1, len(next(iter(pi_results.values()))) + 1))
    for row_idx, trial_number in enumerate(trial_numbers, start=7):
        sheet.cell(row=row_idx, column=2, value=trial_number)

    # Fill in Pi estimates in columns C-F
    for col_idx, count in enumerate(trial_counts, start=3):
        pi_values = pi_results[count]
        for row_idx, pi_value in enumerate(pi_values, start=7):
            sheet.cell(row=row_idx, column=col_idx, value=round(pi_value, 6))

    
    columns = ['C', 'D', 'E', 'F']  # Corresponding to trial_counts in C, D, E, F
    for col_idx, column_letter in enumerate(columns, start=9):  # Columns I to L correspond to 9 to 12
        count = len(pi_results[trial_counts[col_idx - 9]])
        mean_formula = f"=AVERAGE({column_letter}7:{column_letter}{6 + count})"  # Mean formula
        mode_formula = f"=MODE({column_letter}7:{column_letter}{6 + count})"    # Mode formula

        sheet.cell(row=7, column=col_idx, value=mean_formula)  # Insert mean formula in row 7 (I7:L7)
        sheet.cell(row=8, column=col_idx, value=mode_formula)  # Insert mode formula in row 8 (I8:L8)


    update_excel_file_probability(sheet, probability_list)

    
    # Save the workbook
    book.save(file_path)
    print(f"Data in sheet '{sheet_name}' has been replaced successfully.")


def update_excel_file_probability(sheet,probability_list):
    # --- Remove Previous Records from Excel Sheet 
    start_col = 'H'
    end_col = 'M'

    # Clear cells from start_row to the last row in the sheet, from start_col to end_col
    for row in range(18, sheet.max_row + 1):
        for col in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
            cell = sheet.cell(row=row, column=col)
            if not any(cell.coordinate in merged_cell for merged_cell in sheet.merged_cells.ranges):
                cell.value = None


    # Find the maximum round value
    max_round = max(item['Round'] for item in probability_list)

    # Loop through each round and update the cells
    for loopRound in range(max_round):
        rowNumber = loopRound * 3 + 18  # Starting from I15
        columnLetter = 'I'
        columnLetter2 = 'H'

        # Set the round value
        sheet[f'{columnLetter2}{rowNumber}'] = loopRound + 1
        sheet[f'{columnLetter}{rowNumber}'] = 'A'
        sheet[f'{columnLetter}{rowNumber + 1}'] = 'B'
        sheet[f'{columnLetter}{rowNumber + 2}'] = 'C'

        # Merge cells for the round value
        sheet.merge_cells(f'{columnLetter2}{rowNumber}:{columnLetter2}{rowNumber + 2}')

    # Loop through the probability list and update the cells
    for item in probability_list:
        round_number = item['Round']
        rowNumber = (round_number - 1) * 3 + 18  # Adjusting for 0-based index
        columnLetter = 'H'

        if item['Trial Count'] == 1000:
            columnLetter = 'J'
            sheet[f'{columnLetter}{rowNumber}'] = item['Probability Circle']
            sheet[f'{columnLetter}{rowNumber + 1}'] = item['Probability Square']
            sheet[f'{columnLetter}{rowNumber + 2}'] = item['Probability Union']
        elif item['Trial Count'] == 10000:
            columnLetter = 'K'
            sheet[f'{columnLetter}{rowNumber}'] = item['Probability Circle']
            sheet[f'{columnLetter}{rowNumber + 1}'] = item['Probability Square']
            sheet[f'{columnLetter}{rowNumber + 2}'] = item['Probability Union']
        elif item['Trial Count'] == 100000:
            columnLetter = 'L'
            sheet[f'{columnLetter}{rowNumber}'] = item['Probability Circle']
            sheet[f'{columnLetter}{rowNumber + 1}'] = item['Probability Square']
            sheet[f'{columnLetter}{rowNumber + 2}'] = item['Probability Union']
        elif item['Trial Count'] == 1000000:
            columnLetter = 'M'
            sheet[f'{columnLetter}{rowNumber}'] = item['Probability Circle']
            sheet[f'{columnLetter}{rowNumber + 1}'] = item['Probability Square']
            sheet[f'{columnLetter}{rowNumber + 2}'] = item['Probability Union']



# Call Main() to run the script
if __name__ == "__main__":
    Monte_Carlo_Simulation()
