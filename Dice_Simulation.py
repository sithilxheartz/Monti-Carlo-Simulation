from itertools import product
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os  # Required for file operations


def dice_simulation_Main():
    calculate_exact_probability()
    calculate_simulated_probability()

def calculate_exact_probability():

    target_sum = 30
    num_dice = 10
    sides = 6

    # Generate all possible combinations of dice rolls
    all_combinations = product(range(1, sides + 1), repeat=num_dice)

    # Count the number of combinations where the sum of rolls equals the target sum
    favorable_outcomes = sum(1 for combo in all_combinations if sum(combo) == target_sum)

    # Calculate the total number of possible outcomes
    total_outcomes = sides ** num_dice

    exact_probability = favorable_outcomes / total_outcomes

    print(f"Exact probability of obtaining a sum of 30 when rolling 10 dice: {exact_probability}")

def calculate_simulated_probability ():

    trial_count = int(input("How many trials do you want to run? "))

    target_sum = 30
    num_dice = 10
    sides = 6

    target_sum_count = 0

    simulation_results = [] # Stores detailed results for each trial 

    sums_of_simulation = []

    for trial in range(trial_count):

        sum_of_each_attempt = 0

        result = 0

        for _ in range(num_dice):
            side_value = np.random.randint(1, sides + 1)
            sum_of_each_attempt += side_value

        if sum_of_each_attempt == target_sum:
            target_sum_count += 1
            result = True
        else:
            result = False

        simulation_results.append({
            'Trial': trial + 1,
            'Summation': sum_of_each_attempt,
            'Result': result
        })

        sums_of_simulation.append(sum_of_each_attempt)

    # Calculate the simulated probability
    simulated_probability = target_sum_count / trial_count

    print(f"Simulated probability of achieving a sum of 30 by rolling 10 dice over {trial_count} trials: {simulated_probability}")

    # Create a bar chart for the frequency of sums from the simulation
    plt.figure(figsize=(10, 6))
    plt.hist(sums_of_simulation, bins=range(20, 61), edgecolor='black', align='left')
    plt.title('Frequency of Different Sums from 10 Dice Rolls')
    plt.xlabel('Sum of Rolls')
    plt.ylabel('Frequency')
    plt.xticks(range(20, 61))  # Set x-ticks from 20 to 60
    plt.grid(True)
    plt.tight_layout()

    # Save the bar chart as an image
    output_image_path = "./dice_simulation_barchart.png"
    if os.path.exists(output_image_path):
        os.remove(output_image_path)  # Delete the existing file
    plt.savefig(output_image_path)
    print(f"Bar Chart saved to {output_image_path}")

     # =======================================================================================================================

    # Save the workbook
    output_excel_path = "./Coursework.xlsx"

    # Check if the file exists
    if os.path.exists(output_excel_path):
        # Load the existing workbook
        wb = load_workbook(output_excel_path)
        # Check if the sheet already exists
        if "Dice Simulation" in wb.sheetnames:
            # Remove the existing sheet
            del wb["Dice Simulation"]
    else:
        # Create a new workbook
        wb = Workbook()

    # Create a new sheet
    ws = wb.create_sheet("Dice Simulation")

    # Define styles
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write the header row
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)
    cell = ws.cell(row=1, column=1, value="Statistics - Dice Simulation Results")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    cell.font = Font(bold=True, color="FFFFFF", size=15)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Write the headers
    headers = ["Trial", "Summation","Result"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
        ws.row_dimensions[5].height = 25

    # Adjust column widths
    column_widths = [10, 12, 20, 5, 5, 5, 40, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Populate trial data
    for row_num, result in enumerate(simulation_results, 6):
        ws.cell(row=row_num, column=1, value=result['Trial']).border = thin_border
        ws.cell(row=row_num, column=2, value=result['Summation']).border = thin_border
        ws.cell(row=row_num, column=3, value=result['Result']).border = thin_border

    # Write the statistical data section
    ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=8)
    ws.row_dimensions[5].height = 25
    cell = ws.cell(row=5, column=7, value="Statistical Values")
    cell.font = bold_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    stat_headers = [
        "Trial Count",
        "Count of Obtaining a Sum = 30",
        "Probability of Obtaining a Sum = 30", 
    ]

    for row_num, header in enumerate(stat_headers, 6):
        cell = ws.cell(row=row_num, column=7, value=header)
        #cell.font = bold_font
        #cell.fill = header_fill
        cell.border = thin_border

    stat_values = [
        trial_count,
        target_sum_count,
        simulated_probability,
    ]

    for row_num, value in enumerate(stat_values, 6):
        cell = ws.cell(row=row_num, column=8, value=value)
        cell.border = thin_border

    # Save the workbook after all data is written
    wb.save(output_excel_path)
    print(f"Excel file saved to {output_excel_path}")


if __name__ == "__main__":
    calculate_exact_probability()
    calculate_simulated_probability ()