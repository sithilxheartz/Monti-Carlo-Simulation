import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import matplotlib.pyplot as plt
import os  # Required for file operations

def familySimulation_Main():
    family_simulation()
    
def family_simulation():

    trial_count = int(input("How many trials do you want to run? "))

    at_least_one_girl = 0  # Counter for trials with at least one girl
    all_girls = 0  # Counter for trials with all children being girls

    simulation_results = []  # Stores detailed results for each trial 

    for trial in range(trial_count):
        children = [random.choice(['B', 'G']) for _ in range(3)]  # Generate children genders
        girl_count = children.count('G')  # Count the number of girls
        
        is_at_least_one_girl = 'G' in children
        is_all_girls = children == ['G', 'G', 'G']

        # Update counters based on conditions
        if is_at_least_one_girl:
            at_least_one_girl += 1
        if is_all_girls:
            all_girls += 1

        # Store trial results
        simulation_results.append({
            'Trial': trial + 1,
            'Girl Count': girl_count,
            'At least one girl': is_at_least_one_girl,
            'All children are girls': is_all_girls
        })

    # Calculate probabilities
    probability_at_least_one_girl = at_least_one_girl / trial_count
    probability_all_girls = all_girls / trial_count
    conditional_probability = all_girls / at_least_one_girl if at_least_one_girl > 0 else 0

    print(f"The probability of all girls given at least one girl is {conditional_probability:.4f}")

    # Prepare data for scatter plot
    cumulative_at_least_one_girl = [sum([r['At least one girl'] for r in simulation_results[:i+1]]) / (i+1) for i in range(trial_count)]
    cumulative_all_girls = [sum([r['All children are girls'] for r in simulation_results[:i+1]]) / (i+1) for i in range(trial_count)]
    cumulative_conditional = [
        cumulative_all_girls[i] / cumulative_at_least_one_girl[i] if cumulative_at_least_one_girl[i] > 0 else 0
        for i in range(trial_count)
    ]

    # Scatter plot for probabilities
    plt.figure(figsize=(10, 6))
    plt.scatter(range(1, trial_count + 1),
        cumulative_at_least_one_girl, label='Pr(At Least One Girl)', color='blue', alpha=0.7)
    plt.scatter(range(1, trial_count + 1),
        cumulative_all_girls, label='Pr(All Girls)', color='red', alpha=0.7)
    plt.scatter(range(1, trial_count + 1),
        cumulative_conditional, label='Pr(All Girls | At Least One Girl)', color='green', alpha=0.7)

    # Add labels, title, and legend
    plt.title('Scatter Plot of Probabilities Over Trials')
    plt.xlabel('Trial Count')
    plt.ylabel('Probability')
    plt.legend()
    plt.tight_layout()

    # Save the plot as an image
    output_image_path = "./family_simulation_scatterplot.png"
    if os.path.exists(output_image_path):
        os.remove(output_image_path)  # Delete the existing file
    plt.savefig(output_image_path)
    print(f"Scatter Plot saved to {output_image_path}")

    # =======================================================================================================================

    # Save the workbook
    output_excel_path = "./Coursework.xlsx"

    # Check if the file exists
    if os.path.exists(output_excel_path):
        # Load the existing workbook
        wb = load_workbook(output_excel_path)
        # Check if the sheet already exists
        if "Family Simulation" in wb.sheetnames:
            # Remove the existing sheet
            del wb["Family Simulation"]
    else:
        # Create a new workbook
        wb = Workbook()

    # Create a new sheet
    ws = wb.create_sheet("Family Simulation")


    # Define styles
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write the header row
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=8)
    cell = ws.cell(row=1, column=1, value="Statistics - Family Simulation Results")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    cell.font = Font(bold=True, color="FFFFFF", size=15)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    # Write the headers
    headers = ["Trial", "Girl Count", "At least one girl", "All children are girls"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

     # Adjust column widths
    column_widths = [10, 12, 20, 20, 5, 5, 30, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Populate trial data
    for row_num, result in enumerate(simulation_results, 6):
        ws.cell(row=row_num, column=1, value=result['Trial']).border = thin_border
        ws.cell(row=row_num, column=2, value=result['Girl Count']).border = thin_border
        ws.cell(row=row_num, column=3, value=result['At least one girl']).border = thin_border
        ws.cell(row=row_num, column=4, value=result['All children are girls']).border = thin_border

    # Write the statistical data section
    ws.merge_cells(start_row=5, start_column=7, end_row=5, end_column=8)
    ws.row_dimensions[5].height = 25
    cell = ws.cell(row=5, column=7, value="Statistical Values")
    cell.font = bold_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

    stat_headers = [
        "Trial Count",
        "Probability of At Least One Girl",
        "Probability of All Girls", 
        "Conditional Probability"
    ]

    for row_num, header in enumerate(stat_headers, 6):
        cell = ws.cell(row=row_num, column=7, value=header)
        cell.border = thin_border

    values = [
        trial_count,
        probability_at_least_one_girl,
        probability_all_girls, 
        conditional_probability,
    ]

    for row_num, value in enumerate(values, 6):
        cell = ws.cell(row=row_num, column=8, value=value)
        cell.border = thin_border

    # Save the workbook after all data is written
    wb.save(output_excel_path)
    print(f"Excel file saved to {output_excel_path}")


if __name__ == '__main__':
    family_simulation()